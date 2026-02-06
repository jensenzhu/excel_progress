import pandas as pd
from typing import List, Optional, Dict, Any
from pathlib import Path
import io


class ExcelHandler:
    """
    Excel文件处理工具
    负责读取、写入、验证Excel文件
    """
    
    @staticmethod
    def read_excel(file_path: str, sheet_name: Optional[str] = None, 
                   header: int = 0) -> Optional[pd.DataFrame]:
        """
        读取Excel文件
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称，如果为None则读取第一个工作表
            header: 标题行索引
            
        Returns:
            DataFrame对象，如果读取失败返回None
        """
        try:
            if sheet_name is None:
                df = pd.read_excel(file_path, header=header)
            else:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header)
            return df
        except Exception as e:
            print(f"读取Excel文件失败: {e}")
            return None
    
    @staticmethod
    def read_excel_from_bytes(file_bytes: bytes, sheet_name: Optional[str] = None,
                               header: int = 0) -> Optional[pd.DataFrame]:
        """
        从字节流读取Excel文件
        
        Args:
            file_bytes: Excel文件字节流
            sheet_name: 工作表名称，如果为None则读取第一个工作表
            header: 标题行索引
            
        Returns:
            DataFrame对象，如果读取失败返回None
        """
        try:
            file_obj = io.BytesIO(file_bytes)
            if sheet_name is None:
                df = pd.read_excel(file_obj, header=header)
            else:
                df = pd.read_excel(file_obj, sheet_name=sheet_name, header=header)
            return df
        except Exception as e:
            print(f"从字节流读取Excel文件失败: {e}")
            return None
    
    @staticmethod
    def write_excel(df: pd.DataFrame, output_path: str, 
                    sheet_name: str = 'Sheet1', index: bool = False) -> bool:
        """
        写入Excel文件
        
        Args:
            df: DataFrame对象
            output_path: 输出文件路径
            sheet_name: 工作表名称
            index: 是否写入索引
            
        Returns:
            写入成功返回True，否则返回False
        """
        try:
            df.to_excel(output_path, sheet_name=sheet_name, index=index, engine='openpyxl')
            return True
        except Exception as e:
            print(f"写入Excel文件失败: {e}")
            return False
    
    @staticmethod
    def write_excel_to_bytes(df: pd.DataFrame, sheet_name: str = 'Sheet1', 
                             index: bool = False) -> Optional[bytes]:
        """
        将DataFrame写入字节流
        
        Args:
            df: DataFrame对象
            sheet_name: 工作表名称
            index: 是否写入索引
            
        Returns:
            字节流，如果写入失败返回None
        """
        try:
            output = io.BytesIO()
            df.to_excel(output, sheet_name=sheet_name, index=index, engine='openpyxl')
            output.seek(0)
            return output.getvalue()
        except Exception as e:
            print(f"写入字节流失败: {e}")
            return None
    
    @staticmethod
    def get_sheet_names(file_path: str) -> Optional[List[str]]:
        """
        获取Excel文件中的所有工作表名称
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            工作表名称列表，如果获取失败返回None
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            print(f"获取工作表名称失败: {e}")
            return None
    
    @staticmethod
    def get_sheet_names_from_bytes(file_bytes: bytes) -> Optional[List[str]]:
        """
        从字节流获取Excel文件中的所有工作表名称
        
        Args:
            file_bytes: Excel文件字节流
            
        Returns:
            工作表名称列表，如果获取失败返回None
        """
        try:
            import openpyxl
            file_obj = io.BytesIO(file_bytes)
            wb = openpyxl.load_workbook(file_obj, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            print(f"从字节流获取工作表名称失败: {e}")
            return None
    
    @staticmethod
    def validate_excel_file(file_path: str) -> bool:
        """
        验证Excel文件是否有效
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            如果文件有效返回True，否则返回False
        """
        try:
            path = Path(file_path)
            if not path.exists():
                return False
            
            if path.suffix.lower() not in ['.xlsx', '.xls']:
                return False
            
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            wb.close()
            return True
        except Exception as e:
            print(f"验证Excel文件失败: {e}")
            return False
    
    @staticmethod
    def get_file_info(file_path: str) -> Optional[Dict[str, Any]]:
        """
        获取Excel文件信息
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            文件信息字典
        """
        try:
            path = Path(file_path)
            if not path.exists():
                return None
            
            sheet_names = ExcelHandler.get_sheet_names(file_path)
            if not sheet_names:
                return None
            
            first_sheet_df = ExcelHandler.read_excel(file_path, sheet_names[0])
            if first_sheet_df is None:
                return None
            
            return {
                'file_name': path.name,
                'file_path': str(path.absolute()),
                'file_size': path.stat().st_size,
                'sheet_count': len(sheet_names),
                'sheet_names': sheet_names,
                'first_sheet_rows': len(first_sheet_df),
                'first_sheet_columns': len(first_sheet_df.columns),
                'first_sheet_columns_list': first_sheet_df.columns.tolist()
            }
        except Exception as e:
            print(f"获取文件信息失败: {e}")
            return None
    
    @staticmethod
    def create_sample_excel(output_path: str, rows: int = 100, columns: int = 5) -> bool:
        """
        创建示例Excel文件
        
        Args:
            output_path: 输出文件路径
            rows: 行数
            columns: 列数
            
        Returns:
            创建成功返回True，否则返回False
        """
        try:
            import numpy as np
            
            data = {}
            for i in range(columns):
                col_name = f'列{chr(65 + i)}'
                if i % 2 == 0:
                    data[col_name] = np.random.randint(1, 1000, rows)
                else:
                    data[col_name] = [f'数据{j}' for j in range(rows)]
            
            df = pd.DataFrame(data)
            return ExcelHandler.write_excel(df, output_path)
        except Exception as e:
            print(f"创建示例Excel文件失败: {e}")
            return False
    
    @staticmethod
    def merge_excel_files(file_paths: List[str], output_path: str, 
                          sheet_name: str = 'Merged') -> bool:
        """
        合并多个Excel文件
        
        Args:
            file_paths: Excel文件路径列表
            output_path: 输出文件路径
            sheet_name: 输出工作表名称
            
        Returns:
            合并成功返回True，否则返回False
        """
        try:
            all_dfs = []
            for file_path in file_paths:
                df = ExcelHandler.read_excel(file_path)
                if df is not None:
                    all_dfs.append(df)
            
            if not all_dfs:
                return False
            
            merged_df = pd.concat(all_dfs, ignore_index=True)
            return ExcelHandler.write_excel(merged_df, output_path, sheet_name)
        except Exception as e:
            print(f"合并Excel文件失败: {e}")
            return False